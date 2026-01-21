from __future__ import annotations

from datetime import datetime
from decimal import Decimal
from io import BytesIO
from typing import Iterable, Optional

from .pdf_documentos import DocumentoReporteRow


def _safe_str(value) -> str:
    return str(value).strip() if value is not None else ""


def _to_decimal(value) -> Decimal:
    try:
        return Decimal(str(value or 0))
    except Exception:
        return Decimal("0")


def _money(value) -> float:
    try:
        return float(_to_decimal(value))
    except Exception:
        return 0.0


def generar_xlsx_listado_documentos(
    *,
    empresa_nombre: str,
    empresa_ruc: str,
    usuario_nombre: str,
    filtros,  # ReporteFacturacionFiltros (reusado)
    filas: Iterable[DocumentoReporteRow],
    now: Optional[datetime] = None,
) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font

    now_dt = now or datetime.now()

    wb = Workbook()
    ws = wb.active
    ws.title = "Documentos"

    desde = getattr(filtros, "desde", None)
    hasta = getattr(filtros, "hasta", None)
    desde_str = desde.strftime("%d/%m/%Y") if hasattr(desde, "strftime") else ""
    hasta_str = hasta.strftime("%d/%m/%Y") if hasattr(hasta, "strftime") else ""

    ws.append([_safe_str(empresa_nombre).upper()])
    ws.append([f"RUC: {_safe_str(empresa_ruc)}"])
    ws.append([f"Desde: {desde_str}  Hasta: {hasta_str}  Hora: {now_dt.strftime('%H:%M:%S')}  Usuario: {_safe_str(usuario_nombre)}"])
    ws.append([])

    header = ["Tipo", "Número", "Fecha", "CI/RUC", "Nombre", "Total", "Estado"]
    ws.append(header)

    bold = Font(bold=True)
    for cell in ws[5]:
        cell.font = bold
        cell.alignment = Alignment(horizontal="center", vertical="center")

    agrupado_por = _safe_str(getattr(filtros, "agrupado_por", "")).lower()

    current_group = None
    for r in list(filas):
        if agrupado_por == 'clientes':
            grp = (_safe_str(r.nombre) or '(Sin nombre)').strip()
            if grp != current_group:
                current_group = grp
                ws.append([f"CLIENTE/PROVEEDOR: {grp}"])
                row_idx = ws.max_row
                ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=len(header))
                cell = ws.cell(row=row_idx, column=1)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="left", vertical="center")
        fecha_str = r.fecha.strftime("%d/%m/%Y") if hasattr(r.fecha, "strftime") else ""
        ws.append(
            [
                _safe_str(r.tipo),
                _safe_str(r.numero),
                fecha_str,
                _safe_str(r.identificacion),
                _safe_str(r.nombre),
                _money(r.total) if r.total is not None else "",
                _safe_str(r.estado),
            ]
        )

    # Ajustes básicos
    widths = [12, 22, 12, 18, 42, 12, 16]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=5, column=i).column_letter].width = w

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
