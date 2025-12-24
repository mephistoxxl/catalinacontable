from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal
from io import BytesIO
from typing import Iterable, Optional


def _safe_str(value) -> str:
    return str(value).strip() if value is not None else ""


def _to_decimal(value) -> Decimal:
    try:
        return Decimal(str(value or 0))
    except Exception:
        return Decimal("0")


def _money(value) -> float:
    # Excel expects numeric types for money columns.
    try:
        return float(_to_decimal(value))
    except Exception:
        return 0.0


def _extract_tax_vals(fac):
    sub = _to_decimal(getattr(fac, "sub_monto", 0))
    desc = _to_decimal(getattr(fac, "total_descuento", 0))
    prop = _to_decimal(getattr(fac, "propina", 0))
    tot = _to_decimal(getattr(fac, "monto_general", 0))

    sub0 = Decimal("0")
    sub12 = Decimal("0")
    iva_val = Decimal("0")

    try:
        sub0 = _to_decimal(getattr(fac, "subtotal_0", 0))
        sub12 = _to_decimal(getattr(fac, "subtotal_12", 0))
        iva_val = _to_decimal(getattr(fac, "iva_12", 0))
    except Exception:
        pass

    if iva_val == 0:
        try:
            for ti in getattr(fac, "totales_impuestos").all():
                if getattr(ti, "codigo", "") == "2":
                    iva_val += _to_decimal(getattr(ti, "valor", 0))
        except Exception:
            pass

    return sub, desc, sub0, sub12, iva_val, prop, tot


def _fp_label_for(fac) -> str:
    try:
        fps = list(getattr(fac, "formas_pago").all())
        if len(fps) == 1:
            return fps[0].get_forma_pago_display() if hasattr(fps[0], "get_forma_pago_display") else _safe_str(getattr(fps[0], "forma_pago", ""))
        if len(fps) > 1:
            return "Varios"
    except Exception:
        return ""
    return ""


def generar_xlsx_listado_facturacion(
    *,
    empresa_nombre: str,
    empresa_ruc: str,
    usuario_nombre: str,
    filtros,  # ReporteFacturacionFiltros
    facturas: Iterable,
    now: Optional[datetime] = None,
) -> bytes:
    """Genera un XLSX del listado de facturación (mismo dataset filtrado que el PDF)."""

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font

    now_dt = now or datetime.now()

    wb = Workbook()
    ws = wb.active
    ws.title = "Facturacion"

    # Encabezado
    ws.append([_safe_str(empresa_nombre).upper()])
    ws.append([f"RUC: {empresa_ruc}"])
    ws.append([f"Desde: {filtros.desde.strftime('%d/%m/%Y')}  Hasta: {filtros.hasta.strftime('%d/%m/%Y')}  Hora: {now_dt.strftime('%H:%M:%S')}  Usuario: {usuario_nombre}"])
    ws.append(["REPORTE FILTRADO POR:", "Almacenes", ", ".join(getattr(filtros, "almacenes", []) or []), "Formas de pago", ", ".join(getattr(filtros, "formas_pago", []) or []), "Facturadores", ", ".join(getattr(filtros, "facturadores", []) or [])])
    ws.append([])

    header = [
        "Secuencias",
        "Facturas Electrónicas",
        "CI/RUC",
        "Cliente",
        "FP",
        "Emisión",
        "SubTotal",
        "Descuento",
        "SUB.IVA 0%",
        "SUB.IVA 12%",
        "IVA",
        "Propina",
        "TOTAL",
    ]
    ws.append(header)

    bold = Font(bold=True)
    for cell in ws[6]:
        cell.font = bold
        cell.alignment = Alignment(horizontal="center", vertical="center")

    total_sub = Decimal("0")
    total_desc = Decimal("0")
    total_sub0 = Decimal("0")
    total_sub12 = Decimal("0")
    total_iva = Decimal("0")
    total_propina = Decimal("0")
    total_total = Decimal("0")

    for fac in list(facturas):
        sec = f"{_safe_str(getattr(fac, 'establecimiento_formatted', ''))} {_safe_str(getattr(fac, 'punto_emision_formatted', ''))}".strip()
        num = _safe_str(getattr(fac, "secuencia_formatted", ""))
        ci = _safe_str(getattr(fac, "identificacion_cliente", ""))
        cliente = _safe_str(getattr(fac, "nombre_cliente", ""))
        fp = _fp_label_for(fac)
        fecha_em = getattr(fac, "fecha_emision", None)
        fecha_str = fecha_em.strftime("%d/%m/%Y") if hasattr(fecha_em, "strftime") else ""

        sub, desc, sub0, sub12, iva_val, prop, tot = _extract_tax_vals(fac)

        total_sub += sub
        total_desc += desc
        total_sub0 += sub0
        total_sub12 += sub12
        total_iva += iva_val
        total_propina += prop
        total_total += tot

        ws.append(
            [
                sec,
                num,
                ci,
                cliente,
                fp,
                fecha_str,
                _money(sub),
                _money(desc),
                _money(sub0),
                _money(sub12),
                _money(iva_val),
                _money(prop),
                _money(tot),
            ]
        )

    ws.append(
        [
            "",
            "",
            "",
            "TOTAL",
            "",
            "",
            _money(total_sub),
            _money(total_desc),
            _money(total_sub0),
            _money(total_sub12),
            _money(total_iva),
            _money(total_propina),
            _money(total_total),
        ]
    )
    last_row = ws.max_row
    for cell in ws[last_row]:
        cell.font = bold

    # Ajustes básicos
    for col in range(1, len(header) + 1):
        ws.column_dimensions[ws.cell(row=6, column=col).column_letter].width = 16

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
