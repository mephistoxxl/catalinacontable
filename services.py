import io
from typing import List, Tuple, Dict

def parse_product_file(file_obj) -> Tuple[List[Dict], List[str]]:
    """Parse an uploaded XLSX or CSV file returning list of product dicts and errors.

    Nueva especificación de columnas (sin ProductoID):
        Código | Descripción | Barras | Iva | Costo Actual | Precio 1 | Precio 2

    Variantes aceptadas (case-insensitive, ignora espacios/acentos):
        codigo / codigoproducto / codigo_producto
        descripcion / descripcionp
        barras / codigobarras / codigo_barras
        iva
        costoactual / costo_actual / costo
        precio1 / precio 1 / precio
        precio2 / precio 2

    También se mantiene compatibilidad con el formato anterior:
        ProductoID | CodigoProducto | CodigoBarras | DescripcionP | Precio1 | Precio2
    (Si aparece productoid se usará únicamente para lectura, pero ya no es requerido.)

    Retorna (rows, errors). Cada row contiene claves:
        codigo, codigo_barras, descripcion, iva, costo_actual, precio1, precio2
        (y opcionalmente productoid si venía en el archivo legacy)
    """
    import csv
    errors: List[str] = []
    rows: List[Dict] = []

    # Read raw bytes to determine format
    initial = file_obj.read()
    file_obj.seek(0)

    import unicodedata, re

    def _strip_accents(txt: str) -> str:
        return ''.join(c for c in unicodedata.normalize('NFKD', txt) if not unicodedata.combining(c))

    def normalize_header(h: str) -> str:
        h2 = _strip_accents(h or '')
        h2 = h2.lower().strip()
        h2 = re.sub(r'\s+', '', h2)  # remove spaces
        h2 = h2.replace('-', '').replace('_', '')
        return h2

    # Map canonical internal keys to accepted normalized header variants
    header_variants = {
        'productoid': {'productoid','id','idproducto'},  # legacy, opcional
        'codigo': {'codigo','codigoproducto','codigop'},
        'descripcion': {'descripcion','descripcionp','descrip'},
        'codigo_barras': {'barras','codigobarras','codigobarra','codigo_barras','barras13','barras12'},
        'iva': {'iva'},
        'costo_actual': {'costoactual','costo','costo_act','costoactualizado','costo_producto'},
        'precio1': {'precio1','precio','precio_1','precioa'},
        'precio2': {'precio2','precio_2','preciob'},
    }

    # Required new-format keys (precio2 can be optional header but recommended). We will treat precio2 as optional header.
    required_keys = {'codigo','descripcion','codigo_barras','iva','costo_actual','precio1'}

    def process_records(records):
        if not records:
            errors.append('Archivo vacío.')
            return
        raw_headers = records[0]
        normalized_headers = [normalize_header(h) for h in raw_headers]

        # Build mapping internal_key -> column index
        col_map: Dict[str,int] = {}
        for idx, norm in enumerate(normalized_headers):
            for internal, variants in header_variants.items():
                if norm in variants and internal not in col_map:
                    col_map[internal] = idx
                    break

        missing_new = sorted(list(required_keys - set(k for k in col_map.keys() if k != 'productoid')))
        if missing_new:
            errors.append('Faltan columnas requeridas: ' + ', '.join(missing_new))
            return

        for r in records[1:]:
            # Saltar filas totalmente vacías
            if not any(str(c).strip() for c in r):
                continue
            try:
                def safe_get(key):
                    idx = col_map.get(key)
                    if idx is None or idx >= len(r):
                        return ''
                    val = r[idx]
                    return '' if val is None else str(val).strip()

                raw_iva = safe_get('iva') or '0'
                raw_iva_norm = raw_iva.replace('%','').strip()
                # Mantener tal cual; la vista decidirá cómo mapear al modelo.

                data = {
                    'codigo': safe_get('codigo'),
                    'codigo_barras': safe_get('codigo_barras'),
                    'descripcion': safe_get('descripcion'),
                    'iva': raw_iva_norm,
                    'costo_actual': safe_get('costo_actual'),
                    'precio1': safe_get('precio1'),
                    'precio2': safe_get('precio2'),
                }
                if 'productoid' in col_map:
                    data['productoid'] = safe_get('productoid')
            except Exception as e:
                errors.append(f"Error leyendo fila: {e}")
                continue
            # Validación mínima
            if not data['codigo']:
                errors.append('Fila omitida: código vacío.')
                continue
            rows.append(data)

    # Try XLSX first
    try:
        from openpyxl import load_workbook  # type: ignore
        bio = io.BytesIO(initial)
        wb = load_workbook(bio, data_only=True)
        ws = wb.active
        records = []
        for row in ws.iter_rows(values_only=True):
            records.append([str(c) if c is not None else '' for c in row])
        if records:
            process_records(records)
            return rows, errors
    except ModuleNotFoundError:
        # openpyxl not installed; fallback to CSV attempt below
        pass
    except Exception:
        # Not a valid xlsx; attempt CSV
        pass

    # CSV fallback
    try:
        text = initial.decode('utf-8', errors='ignore')
        reader = csv.reader(io.StringIO(text))
        records = list(reader)
        if records:
            process_records(records)
    except Exception as e:
        errors.append(f"Error procesando archivo CSV: {e}")
    finally:
        file_obj.seek(0)
    return rows, errors
import requests
import os
from dotenv import load_dotenv
import json
import logging
from typing import Any, Dict, Iterable, Optional, Union
from requests.exceptions import RequestException, Timeout
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry

# Configurar logging con más detalle
logging.basicConfig(
    level=logging.DEBUG,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

load_dotenv()

DEFAULT_DIRECCION = 'SIN DIRECCIÓN REGISTRADA'


def _limpiar_texto(valor: Any) -> str:
    if valor is None:
        return ''
    if isinstance(valor, str):
        texto = valor.strip()
    else:
        texto = str(valor).strip()
    return texto


def _agregar_componentes(destino: list[str], vistos: set[str], *valores: Any) -> None:
    for valor in valores:
        if not valor:
            continue
        if isinstance(valor, (list, tuple, set)):
            _agregar_componentes(destino, vistos, *valor)
            continue
        texto = _limpiar_texto(valor)
        if not texto:
            continue
        clave = texto.lower()
        if clave in vistos:
            continue
        vistos.add(clave)
        destino.append(texto)


def _unir_componentes(componentes: Iterable[Any], separador: str = ' ') -> str:
    partes = []
    for componente in componentes:
        texto = _limpiar_texto(componente)
        if texto:
            partes.append(texto)
    return separador.join(partes)


def _extraer_direccion(data: Dict[str, Any], tipo_identificacion: str) -> str:
    componentes: list[str] = []
    vistos: set[str] = set()

    def agregar(*items: Any) -> None:
        _agregar_componentes(componentes, vistos, *items)

    if tipo_identificacion == 'RUC':
        establecimientos = data.get('establecimientos') or []
        for establecimiento in establecimientos:
            if not isinstance(establecimiento, dict):
                continue
            agregar(
                establecimiento.get('direccionCompleta'),
                establecimiento.get('direccion'),
                _unir_componentes(
                    [
                        establecimiento.get('calle'),
                        establecimiento.get('numero'),
                        establecimiento.get('interseccion'),
                    ]
                ),
                _unir_componentes(
                    [
                        establecimiento.get('provincia'),
                        establecimiento.get('canton'),
                        establecimiento.get('parroquia'),
                    ],
                    ' - '
                ),
            )
            if componentes:
                break

        agregar(
            data.get('direccionMatriz'),
            _unir_componentes(
                [data.get('calleMatriz'), data.get('numeroMatriz'), data.get('interseccionMatriz')]
            ),
            _unir_componentes(
                [
                    data.get('provinciaMatriz'),
                    data.get('cantonMatriz'),
                    data.get('parroquiaMatriz'),
                ],
                ' - '
            ),
        )

    agregar(
        data.get('direccionDomicilio'),
        data.get('direccion'),
        data.get('direccionCompleta'),
        _unir_componentes(
            [data.get('calleDomicilio'), data.get('numeroDomicilio'), data.get('referenciaDomicilio')]
        ),
        data.get('barrio'),
        data.get('parroquia'),
        data.get('ciudad'),
        data.get('provincia'),
    )

    return ' - '.join(componentes)

def validar_identificacion(identificacion: str) -> Optional[str]:
    """Valida el formato de la identificación.

    Retorna el tipo de identificación (`'RUC'` o `'CEDULA'`) si es válida.
    """
    logger.debug(f"Iniciando validación de identificación: {identificacion}")
    if not identificacion or not isinstance(identificacion, str):
        logger.error(
            f"Identificación inválida: valor vacío o tipo incorrecto - {type(identificacion)}"
        )
        return None
    if not identificacion.isdigit():
        logger.error("Identificación inválida: contiene caracteres no numéricos")
        return None
    if len(identificacion) == 13:
        logger.debug("Identificación válida: RUC")
        return "RUC"
    if len(identificacion) == 10:
        logger.debug("Identificación válida: CÉDULA")
        return "CEDULA"
    logger.error(
        f"Identificación inválida: longitud {len(identificacion)} no permitida"
    )
    return None

def consultar_identificacion(identificacion: str) -> Dict[str, Union[str, bool]]:
    """Consulta información de RUC o cédula usando el API de Zampisoft."""
    logger.info(f"Consultando identificación: {identificacion}")

    tipo_identificacion = validar_identificacion(identificacion)
    if not tipo_identificacion:
        return {
            'error': True,
            'message': 'Identificación inválida: debe tener 10 o 13 dígitos',
            'status_code': 400
        }
    
    # URL de la API de Zampisoft
    url = "https://apiconsult.zampisoft.com/api/consultar"
    
    # Obtener token del archivo .env o usar el token por defecto
    token = os.getenv('ZAMPISOFT_TOKEN', 'wTGv-8Iqi-ckFW-A8bo')
    logger.info(f"Token a usar: {token[:4]}...{token[-4:]}")
    
    try:
        # Configurar los parámetros
        params = {
            "identificacion": identificacion,
            "token": token
        }
        
        # Mostrar información de la solicitud
        logger.info(f"URL de la API: {url}")
        logger.info(f"Parámetros de la solicitud: {json.dumps(params, indent=2)}")
        
        # Realizar la solicitud
        logger.info("Enviando solicitud a la API...")
        response = requests.get(url, params=params, timeout=15)
        
        # Mostrar información de la respuesta
        logger.info(f"Status Code: {response.status_code}")
        logger.info(f"Headers de respuesta: {dict(response.headers)}")
        logger.info(f"URL final: {response.url}")
        logger.info(f"Contenido de la respuesta: {response.text[:1000]}")
        
        # Intentar parsear la respuesta como JSON
        try:
            data = response.json()
            logger.info(f"Respuesta JSON completa: {json.dumps(data, indent=2, ensure_ascii=False)}")
        except json.JSONDecodeError as e:
            logger.error(f"Error al decodificar JSON: {str(e)}")
            logger.error(f"Respuesta recibida: {response.text[:500]}")
            return {
                'error': True,
                'message': 'Error al procesar la respuesta del servidor',
                'status_code': 500
            }
        
        # Verificar si hay error en la respuesta de la API
        api_error = data.get('error', False)
        api_message = data.get('message', 'Error desconocido de la API') if api_error else ''

        # Extraer dirección del primer establecimiento si existe
        establecimientos = data.get('establecimientos', [])
        nombre_comercial = ''
        if tipo_identificacion == 'RUC' and establecimientos:
            primer_establecimiento = establecimientos[0] if isinstance(establecimientos[0], dict) else {}
            nombre_comercial = primer_establecimiento.get('nombreFantasiaComercial', '')

        direccion = _extraer_direccion(data, tipo_identificacion)

        # Mapear la respuesta al formato esperado
        tipo_contribuyente = data.get('tipoContribuyente')
        logger.info(f"Tipo de contribuyente original: {tipo_contribuyente}")

        # Mapear el tipo de contribuyente a los valores permitidos en el modelo.
        # Solo se aplica para consultas de RUC, ya que la API no entrega datos
        # de régimen para consultas con cédula y en esos casos no se debe
        # modificar el valor existente en el formulario.
        tipo_regimen_mapeado = None
        if tipo_contribuyente and tipo_identificacion == 'RUC':
            tipo_regimen_mapeado = (
                'RIMPE' if 'RIMPE' in tipo_contribuyente.upper() else 'GENERAL'
            )

        logger.info(f"Tipo de régimen mapeado: {tipo_regimen_mapeado}")
        
        # Mapear obligado a llevar contabilidad a SI/NO
        obligado_contabilidad = data.get('obligadoLlevarContabilidad', '')
        obligado_mapeado = 'SI' if obligado_contabilidad and obligado_contabilidad.upper() == 'SI' else 'NO'
        
        logger.info(f"Obligado a llevar contabilidad original: {obligado_contabilidad}, mapeado: {obligado_mapeado}")
        
        razon_social = data.get('razonSocial', '')
        if not razon_social:
            nombres = data.get('nombres', data.get('nombre', ''))
            apellidos = data.get('apellidos', data.get('apellido', ''))
            razon_social = f"{nombres} {apellidos}".strip()

        resultado = {
            'error': api_error,
            'message': api_message,
            'razon_social': razon_social,
            'nombre_comercial': nombre_comercial,
            'direccion': direccion or DEFAULT_DIRECCION,
            'telefono': data.get('telefono', ''),
            'email': data.get('email', ''),
            'tipo_contribuyente': tipo_contribuyente,
            'estado': data.get('estadoContribuyenteRuc', data.get('estado', '')),
            'obligado_contabilidad': obligado_mapeado,
            'actividad_economica': data.get('actividadEconomicaPrincipal', ''),
            'status_code': response.status_code,
            'tipo_identificacion': tipo_identificacion
        }

        if tipo_regimen_mapeado:
            resultado['tipo_regimen'] = tipo_regimen_mapeado
        
        logger.info(f"Resultado mapeado: {json.dumps(resultado, indent=2, ensure_ascii=False)}")
        return resultado
        
    except requests.exceptions.RequestException as e:
        logger.error(f"Error de conexión: {str(e)}")
        return {
            'error': True,
            'message': f'Error de conexión: {str(e)}',
            'status_code': 500
        }
    except Exception as e:
        logger.error(f"Error inesperado: {str(e)}")
        import traceback
        logger.error(f"Traceback: {traceback.format_exc()}")
        return {
            'error': True,
            'message': f'Error inesperado: {str(e)}',
            'status_code': 500
        }


# Las funciones de test siguen igual...
def test_consulta_identificacion():
    """Función de prueba para verificar la consulta de identificaciones."""
    # Identificación de prueba (puedes cambiar por una que sepas que existe)
    identificacion_prueba = "1713959011001"

    print(f"\n{'='*60}")
    print(f"PRUEBA DE CONSULTA IDENTIFICACIÓN: {identificacion_prueba}")
    print(f"{'='*60}\n")

    # Validar identificación
    tipo = validar_identificacion(identificacion_prueba)
    print(f"Tipo de identificación: {tipo}")

    if tipo:
        # Realizar consulta
        resultado = consultar_identificacion(identificacion_prueba)
        
        print("\nRESULTADO DE LA CONSULTA:")
        print("-" * 40)
        
        if resultado.get('error'):
            print(f"ERROR: {resultado.get('message')}")
            print(f"Código de estado: {resultado.get('status_code')}")
        else:
            print("CONSULTA EXITOSA")
            print("\nDatos obtenidos:")
            for campo, valor in resultado.items():
                if campo != 'error' and valor:
                    print(f"  {campo}: {valor}")
        
    print(f"\n{'='*60}\n")
    
    return resultado


def debug_api_response():
    """
    Función para debug detallado de la respuesta de la API
    """
    import requests
    
    # Configuración
    url = "https://apiconsult.zampisoft.com/api/consultar"
    params = {
        "identificacion": "1713959011001",  # RUC de prueba
        "token": "7a7R-zcYo-7pB9-hkqN"
    }
    
    print(f"\n{'='*60}")
    print("DEBUG DETALLADO DE LA API")
    print(f"{'='*60}\n")
    
    print(f"URL: {url}")
    print(f"Parámetros: {params}")
    
    try:
        # Hacer la solicitud
        response = requests.get(url, params=params, timeout=15)
        
        print(f"\nRespuesta:")
        print(f"  Status Code: {response.status_code}")
        print(f"  Headers: {dict(response.headers)}")
        print(f"  URL final: {response.url}")
        print(f"\nContenido de la respuesta:")
        print("-" * 40)
        print(response.text)
        print("-" * 40)
        
        # Intentar parsear como JSON
        try:
            data = response.json()
            print("\nDatos parseados como JSON:")
            print(json.dumps(data, indent=2, ensure_ascii=False))
        except:
            print("\nNota: La respuesta no es JSON válido")
            
    except Exception as e:
        print(f"\nERROR: {str(e)}")
        import traceback
        traceback.print_exc()
    
    print(f"\n{'='*60}\n")

def test_api_directo(ruc=None):
    """
    Prueba directa de la API sin usar Django
    Args:
        ruc (str): El RUC a consultar. Si no se proporciona, se usará uno de prueba.
    """
    import requests
    import json
    
    # Si no se proporciona RUC, usar uno de prueba
    if not ruc:
        ruc = "1713959011001"
        print("\n⚠️  Usando RUC de prueba. Para probar con un RUC específico, llame a la función con el RUC como parámetro.")
        print("Ejemplo: test_api_directo('1234567890001')")
    
    # URL y parámetros
    url = "https://apiconsult.zampisoft.com/api/consultar"
    params = {
        "identificacion": ruc,
        "token": "wTGv-8Iqi-ckFW-A8bo"
    }
    
    print("\n" + "="*80)
    print("PRUEBA DIRECTA DE LA API")
    print("="*80)
    print(f"\nConsultando RUC: {ruc}")
    print(f"URL: {url}")
    print(f"Parámetros: {json.dumps(params, indent=2)}")
    
    try:
        # Realizar la solicitud
        response = requests.get(url, params=params, timeout=15)
        print(f"\nStatus Code: {response.status_code}")
        
        # Mostrar la respuesta completa
        print("\n" + "-"*80)
        print("RESPUESTA COMPLETA DE LA API:")
        print("-"*80)
        print(response.text)
        print("-"*80)
        
        # Intentar parsear como JSON
        try:
            data = response.json()
            print("\nDatos parseados como JSON:")
            print(json.dumps(data, indent=2, ensure_ascii=False))
            
            # Mostrar campos específicos
            print("\n" + "-"*80)
            print("CAMPOS ESPECÍFICOS:")
            print("-"*80)
            campos = [
                ('razonSocial', 'Razón Social'),
                ('nombreComercial', 'Nombre Comercial'),
                ('direccion', 'Dirección'),
                ('telefono', 'Teléfono'),
                ('email', 'Email'),
                ('tipoContribuyente', 'Tipo Contribuyente'),
                ('estado', 'Estado'),
                ('obligadoContabilidad', 'Obligado Contabilidad')
            ]
            
            for campo, nombre in campos:
                valor = data.get(campo, 'No encontrado')
                print(f"{nombre}: {valor}")
                
        except json.JSONDecodeError as e:
            print(f"\nError al parsear JSON: {str(e)}")
            print("La respuesta no es un JSON válido")
            
    except Exception as e:
        print(f"\nError en la solicitud: {str(e)}")
    
    print("\n" + "="*80)
    return data if 'data' in locals() else None

if __name__ == "__main__":
    # Ejemplo de uso:
    # test_api_directo()  # Usa RUC de prueba
    # test_api_directo('1234567890001')  # Usa RUC específico
    test_api_directo()
